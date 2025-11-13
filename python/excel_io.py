"""Read and write simple Excel files using :mod:`openpyxl`.

Usage example
-------------
>>> from python import excel_io
>>> excel_io.write_workbook('example.xlsx', [{'name': 'Ada'}])  # doctest: +SKIP
PosixPath('example.xlsx')
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence

try:  # pragma: no cover - optional dependency
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

PathLike = str | Path
Row = dict[str, object]

__all__ = ["ExcelError", "read_workbook", "write_workbook"]


class ExcelError(RuntimeError):
    """Raised when Excel operations fail."""


def _require_openpyxl() -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for Excel operations")


def read_workbook(path: PathLike, *, sheet: str | int | None = 0) -> list[Row]:
    """Return ``sheet`` as a list of dictionaries."""

    _require_openpyxl()
    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
        if sheet is None:
            worksheet = workbook.active
        elif isinstance(sheet, int):
            worksheet = workbook.worksheets[sheet]
        else:
            worksheet = workbook[sheet]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value) if value is not None else f"column_{index}" for index, value in enumerate(rows[0])]
        records: list[Row] = []
        for raw in rows[1:]:
            record = {header: cell for header, cell in zip(headers, raw)}
            records.append(record)
        return records
    except Exception as exc:
        raise ExcelError(f"Failed to read Excel file {path}: {exc}") from exc


def write_workbook(
    path: PathLike,
    rows: Iterable[Row],
    *,
    headers: Sequence[str] | None = None,
    sheet_name: str = "Sheet1",
) -> Path:
    """Write ``rows`` to ``path`` and return the resulting :class:`Path`."""

    _require_openpyxl()
    output = Path(path)
    try:
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        iterator = iter(rows)
        first = next(iterator, None)
        if first is None:
            workbook.save(output)
            return output
        rest = [first, *iterator]
        headers = list(headers or first.keys())
        worksheet.append(list(headers))
        for row in rest:
            worksheet.append([row.get(header) for header in headers])
        workbook.save(output)
        return output
    except Exception as exc:
        raise ExcelError(f"Failed to write Excel file {output}: {exc}") from exc
